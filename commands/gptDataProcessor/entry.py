import adsk.core
import adsk.fusion
import adsk.cam
import traceback
import os
import sys
import math

# Import fusion utilities
try:
    from ...lib import fusionAddInUtils as futil
    from ... import config
except ImportError:
    # Fallback for when running as script
    try:
        import sys
        import os
        current_dir = os.path.dirname(os.path.abspath(__file__))
        parent_dir = os.path.dirname(os.path.dirname(current_dir))
        if parent_dir not in sys.path:
            sys.path.append(parent_dir)
        from lib import fusionAddInUtils as futil
        import config
    except ImportError:
        # Create minimal fallback
        class FakeUtil:
            def log(self, msg): print(f"LOG: {msg}")
            def add_handler(self, *args, **kwargs): pass
        futil = FakeUtil()

        class FakeConfig:
            COMPANY_NAME = 'ACME'
            ADDIN_NAME = 'stent_frame'
        config = FakeConfig()

# Add the parent directory to sys.path to import crown_arc module
parent_dir = os.path.dirname(os.path.dirname(
    os.path.dirname(os.path.abspath(__file__))))
if parent_dir not in sys.path:
    sys.path.append(parent_dir)

# Import the crown_arc module for calculations
try:
    import crown_arc
except ImportError:
    crown_arc = None

# TODO *** Define the location of the command ***
# This is done by declaring the space, the tab, and the panel.
CMD_ID = f'{config.COMPANY_NAME}_{config.ADDIN_NAME}_gptDataProcessor'

# Global variable to store parameters from JSON files
global_parameters = {}
CMD_NAME = 'Process Stent Data'
CMD_Description = 'Load and process CSV, JSON, or Excel file with detailed stent frame data'

# Specify that the command will be promoted to the panel.
IS_PROMOTED = True

# TODO *** Define the location of the command ***
# This is done by declaring the space, the tab, and the panel.
WORKSPACE_ID = 'FusionSolidEnvironment'
# Try the Scripts and Add-ins panel instead
PANEL_ID = 'SolidScriptsAddinsPanel'
COMMAND_BESIDE_ID = ''  # Place at the end instead of beside a specific command

# Resource file folders relative to this file.
ICON_FOLDER = os.path.join(os.path.dirname(
    os.path.abspath(__file__)), 'resources', '')

# Local list of event handlers used to maintain a reference so
# they are not released and garbage collected.
local_handlers = []

# Global variables
excel_file_path = ""


def start():
    # Create a command Definition.
    try:
        print(f'DEBUG: Starting {CMD_NAME} command...')
        futil.log(f'Starting {CMD_NAME} command...')

        cmd_def = adsk.core.Application.get().userInterface.commandDefinitions.addButtonDefinition(
            CMD_ID, CMD_NAME, CMD_Description, ICON_FOLDER)

        print(f'DEBUG: Command definition created: {CMD_ID}')
        futil.log(f'Command definition created: {CMD_ID}')

        # Define an event handler for the command created event. It will be called when the button is clicked.
        futil.add_handler(cmd_def.commandCreated, command_created)

        # ******** Add a button into the UI so the user can run the command. ********
        # Get the target workspace the button will be created in.
        workspace = adsk.core.Application.get(
        ).userInterface.workspaces.itemById(WORKSPACE_ID)
        print(
            f'DEBUG: Got workspace: {workspace.name if workspace else "None"}')
        futil.log(f'Got workspace: {workspace.name if workspace else "None"}')

        # Get the panel the button will be created in.
        panel = workspace.toolbarPanels.itemById(PANEL_ID)
        print(f'DEBUG: Got panel: {panel.name if panel else "None"}')
        futil.log(f'Got panel: {panel.name if panel else "None"}')

        # Create the button command control in the UI after the specified existing command.
        if COMMAND_BESIDE_ID:
            control = panel.controls.addCommand(
                cmd_def, COMMAND_BESIDE_ID, False)
        else:
            control = panel.controls.addCommand(cmd_def)

        print(
            f'DEBUG: Added control to panel: {control.id if control else "None"}')
        futil.log(
            f'Added control to panel: {control.id if control else "None"}')

        # Specify if the command is promoted to the main toolbar.
        control.isPromoted = IS_PROMOTED
        print(f'DEBUG: {CMD_NAME} command started successfully')
        futil.log(f'{CMD_NAME} command started successfully')

        # Debug: List all controls in the panel to see if our button is there
        print(f'DEBUG: Panel controls count: {panel.controls.count}')
        for i in range(panel.controls.count):
            ctrl = panel.controls.item(i)
            print(f'DEBUG: Control {i}: {ctrl.id} - {ctrl.objectType}')

    except Exception as e:
        print(f'ERROR: Error starting {CMD_NAME} command: {str(e)}')
        futil.log(f'Error starting {CMD_NAME} command: {str(e)}')
        import traceback
        print(f'ERROR: Traceback: {traceback.format_exc()}')
        futil.log(traceback.format_exc())


def stop():
    # Get the various UI elements for this command
    workspace = adsk.core.Application.get(
    ).userInterface.workspaces.itemById(WORKSPACE_ID)
    panel = workspace.toolbarPanels.itemById(PANEL_ID)
    command_control = panel.controls.itemById(CMD_ID)
    command_definition = adsk.core.Application.get(
    ).userInterface.commandDefinitions.itemById(CMD_ID)

    # Delete the button command control
    if command_control:
        command_control.deleteMe()

    # Delete the command definition
    if command_definition:
        command_definition.deleteMe()

# Function that is called when a user clicks the corresponding button in the UI.
# This defines the contents of the command dialog and connects to the command related events.


def command_created(args: adsk.core.CommandCreatedEventArgs):
    # General logging for debug.
    futil.log(f'{CMD_NAME} Command Created Event')
    print("DEBUG: Starting command creation...")

    # Set dialog size - make it wider for better layout
    args.command.isOKButtonVisible = True
    args.command.setDialogInitialSize(600, 500)  # width, height in pixels
    args.command.setDialogMinimumSize(500, 400)  # minimum width, height
    print("DEBUG: Dialog size set successfully")

    # https://help.autodesk.com/view/fusion360/ENU/?contextId=CommandInputs
    inputs = args.command.commandInputs
    print("DEBUG: Got command inputs")

    # Create a default value using a string. This method also accepts a real as a parameter.
    # If a real is provided, the value will be assumed to be in current user units.

    # Add a message to describe the command
    message_input = inputs.addTextBoxCommandInput('message', '',
                                                  'This command processes Excel files with detailed stent frame data.\n\n'
                                                  'The Excel file should contain a sheet named "WaveInputsByColumn" with\n'
                                                  'columns: ring, col, wave_height_mm, wave_width_mm, strut_width_mm,\n'
                                                  'gap_above_mm, gap_below_mm, upper_chord_center_mm, upper_sagitta_center_mm,\n'
                                                  'upper_outer_sagitta_mm, lower_chord_center_mm, lower_sagitta_center_mm,\n'
                                                  'lower_outer_sagitta_mm, theta_deg, Rc_mm', 6, True)
    print("DEBUG: Message input added")

    # File selection group
    file_group = inputs.addGroupCommandInput(
        'file_group', 'Excel File Selection')
    file_group.isExpanded = True
    file_group_inputs = file_group.children
    print("DEBUG: File group created")

    # Add a text input for the file path
    file_path_input = file_group_inputs.addStringValueInput(
        'file_path', 'Excel File Path', '')
    file_path_input.tooltip = 'Path to the Excel file containing stent data'
    print("DEBUG: File path input added")

    # Add a button to browse for file
    browse_button = file_group_inputs.addBoolValueInput(
        'browse_file', 'Browse for File', False, '', False)
    browse_button.tooltip = 'Click to browse for Excel file'
    print("DEBUG: Browse button added")

    # Processing options group
    try:
        options_group = inputs.addGroupCommandInput(
            'options_group', 'Processing Options')
        options_group.isExpanded = True
        options_group_inputs = options_group.children
        print("DEBUG: Options group created successfully")
    except Exception as e:
        print(f"ERROR: Failed to create options group: {e}")
        return

    # Add diameter input (will be calculated from data)
    try:
        diameter_input = options_group_inputs.addValueInput('diameter', 'Stent Diameter', 'mm',
                                                            adsk.core.ValueInput.createByReal(0.19))
        diameter_input.tooltip = 'Stent diameter in millimeters (will be read from file if available)'
        print("DEBUG: Diameter input created successfully")
    except Exception as e:
        print(f"ERROR: Failed to create diameter input: {e}")

    # Add length input (will be calculated from data)
    try:
        length_input = options_group_inputs.addValueInput('length', 'Stent Length', 'mm',
                                                          adsk.core.ValueInput.createByReal(0.8))
        length_input.tooltip = 'Stent length in millimeters (will be read from file if available)'
        print("DEBUG: Length input created successfully")
    except Exception as e:
        print(f"ERROR: Failed to create length input: {e}")

    # Add number of rings input (read-only, from file)
    num_rings_input = options_group_inputs.addIntegerSpinnerCommandInput(
        'num_rings', 'Number of Rings', 1, 20, 1, 6)
    num_rings_input.tooltip = 'Number of rings (will be read from file if available)'

    # Add crowns per ring input (read-only, from file)
    crowns_per_ring_input = options_group_inputs.addIntegerSpinnerCommandInput(
        'crowns_per_ring', 'Crowns per Ring', 1, 32, 1, 8)
    crowns_per_ring_input.tooltip = 'Number of crowns per ring (will be read from file if available)'

    # Add option to draw construction lines
    draw_construction = options_group_inputs.addBoolValueInput(
        'draw_construction', 'Draw Construction Lines', True, '', True)
    draw_construction.tooltip = 'Draw construction lines showing wave boundaries and centers'

    # Add option to draw chord lines
    draw_chords = options_group_inputs.addBoolValueInput(
        'draw_chords', 'Draw Chord Lines', True, '', True)
    draw_chords.tooltip = 'Draw chord lines at sagitta positions for each crown'

    # Add option to create sketch points
    create_points = options_group_inputs.addBoolValueInput(
        'create_points', 'Create Sketch Points', True, '', False)
    create_points.tooltip = 'Create sketch points at key intersections'

    # Status group
    status_group = inputs.addGroupCommandInput('status_group', 'Status')
    status_group.isExpanded = False
    status_group_inputs = status_group.children

    # Add status text
    status_text = status_group_inputs.addTextBoxCommandInput(
        'status', '', 'Ready to process Excel file.', 2, True)
    print("DEBUG: Status group added successfully")

    # Connect to the events that are needed by this command.
    futil.add_handler(args.command.execute, command_execute,
                      local_handlers=local_handlers)
    futil.add_handler(args.command.inputChanged,
                      command_input_changed, local_handlers=local_handlers)
    futil.add_handler(args.command.executePreview,
                      command_preview, local_handlers=local_handlers)
    futil.add_handler(args.command.validateInputs,
                      command_validate_input, local_handlers=local_handlers)
    futil.add_handler(args.command.destroy, command_destroy,
                      local_handlers=local_handlers)
    print("DEBUG: Event handlers registered successfully")


# This event handler is called when the user clicks the OK button in the command dialog or
# is immediately called after the created event not command inputs were created for the dialog.


def command_execute(args: adsk.core.CommandEventArgs):
    # General logging for debug.
    futil.log(f'{CMD_NAME} Command Execute Event')

    # TODO *** Do something useful here in the command. ***

    # Get the values from the command inputs.
    inputs = args.command.commandInputs

    file_path_input = adsk.core.StringValueCommandInput.cast(
        inputs.itemById('file_path'))
    diameter_input = adsk.core.ValueCommandInput.cast(
        inputs.itemById('diameter'))
    draw_construction_input = adsk.core.BoolValueCommandInput.cast(
        inputs.itemById('draw_construction'))
    draw_chords_input = adsk.core.BoolValueCommandInput.cast(
        inputs.itemById('draw_chords'))
    create_points_input = adsk.core.BoolValueCommandInput.cast(
        inputs.itemById('create_points'))

    if not file_path_input.value:
        adsk.core.Application.get().userInterface.messageBox(
            'Please select an Excel file first.')
        return

    # Process the Excel file and create the stent frame
    try:
        length_input = adsk.core.ValueCommandInput.cast(
            inputs.itemById('length'))
        process_excel_file(
            file_path=file_path_input.value,
            diameter_mm=diameter_input.value * 10,  # Convert cm to mm
            length_mm=length_input.value * 10 if length_input else None,  # Convert cm to mm
            draw_construction=draw_construction_input.value,
            draw_chords=draw_chords_input.value,
            create_points=create_points_input.value
        )
    except Exception as e:
        adsk.core.Application.get().userInterface.messageBox(
            f'Error processing Excel file: {str(e)}')
        futil.log(f'Error in command_execute: {traceback.format_exc()}')

# This event handler is called when the command needs to compute a new preview in the graphics window.


def command_preview(args: adsk.core.CommandEventArgs):
    # General logging for debug.
    futil.log(f'{CMD_NAME} Command Preview Event')
    inputs = args.command.commandInputs

# This event handler is called when the user changes anything in the command dialog
# allowing you to modify values of other inputs based on that change.


def command_input_changed(args: adsk.core.InputChangedEventArgs):
    changed_input = args.input
    inputs = args.inputs

    futil.log(
        f'{CMD_NAME} Input Changed Event fired from a change to {changed_input.id}')

    # Handle file browse button
    if changed_input.id == 'browse_file':
        browse_button = adsk.core.BoolValueCommandInput.cast(changed_input)
        if browse_button.value:
            # Reset the button
            browse_button.value = False

            # Open file dialog
            file_dialog = adsk.core.Application.get().userInterface.createFileDialog()
            file_dialog.isMultiSelectEnabled = False
            file_dialog.title = "Select JSON, CSV, or Excel File"
            file_dialog.filter = "JSON files (*.json);;CSV files (*.csv);;Excel files (*.xlsx);;All files (*.*)"

            if file_dialog.showOpen() == adsk.core.DialogResults.DialogOK:
                file_path = file_dialog.filename
                file_path_input = adsk.core.StringValueCommandInput.cast(
                    inputs.itemById('file_path'))
                file_path_input.value = file_path

                # Update preview
                update_data_preview(inputs, file_path)
                print(f"DEBUG: File dialog selected: {file_path}")

    # Handle file path changes
    elif changed_input.id == 'file_path':
        file_path_input = adsk.core.StringValueCommandInput.cast(changed_input)
        if file_path_input.value and os.path.exists(file_path_input.value):
            print(f"DEBUG: File path changed to: {file_path_input.value}")
            update_data_preview(inputs, file_path_input.value)


def update_data_preview(inputs, file_path):
    """Update the status text with file information"""
    status_text = None

    try:
        status_text = adsk.core.TextBoxCommandInput.cast(
            inputs.itemById('status'))

        if not os.path.exists(file_path):
            if status_text:
                status_text.text = 'Error: File not found.'
            return

        # Try to read Excel file
        file_data = read_excel_data(file_path)
        data = file_data['data']
        parameters = file_data.get('parameters', {})

        print(f"DEBUG: Read {len(data)} data rows")
        print(f"DEBUG: Parameters found: {parameters}")

        if data is None or len(data) == 0:
            if status_text:
                status_text.text = 'Error: No valid data found.'
            return

        # Create preview text
        rings = sorted(set(row['ring'] for row in data))
        cols_per_ring = {}
        for ring in rings:
            ring_data = [row for row in data if row['ring'] == ring]
            cols_per_ring[ring] = len(ring_data)

        total_rows = len(data)

        preview = f"File loaded successfully!\n\n"
        preview += f"Total data rows: {total_rows}\n"
        preview += f"Rings found: {len(rings)} (rings {min(rings)} to {max(rings)})\n"
        preview += f"Columns per ring: {list(cols_per_ring.values())}\n"

        # Show parameters from file if available
        if parameters:
            preview += f"\nParameters from file:\n"
            if 'diameter_mm' in parameters:
                preview += f"  Diameter: {parameters['diameter_mm']} mm\n"
            if 'length_mm' in parameters:
                preview += f"  Length: {parameters['length_mm']} mm\n"
            if 'num_rings' in parameters:
                preview += f"  Number of rings: {parameters['num_rings']}\n"
            if 'crowns_per_ring' in parameters:
                preview += f"  Crowns per ring: {parameters['crowns_per_ring']}\n"

        # Show first few rows as sample
        preview += "Sample data (first 3 rows):\n"
        for i, row in enumerate(data[:3]):
            preview += f"Ring {row['ring']}, Col {row['col']}: "
            preview += f"height={row['wave_height_mm']:.3f}mm, "
            preview += f"width={row['wave_width_mm']:.3f}mm\n"

        if len(data) > 3:
            preview += f"... and {len(data) - 3} more rows"

        if status_text:
            status_text.text = f'Ready to process {total_rows} data points from {len(rings)} rings.'

        # Store parameters globally for later access
        global global_parameters
        global_parameters = parameters.copy()
        print(f"DEBUG: Stored parameters globally: {global_parameters}")

        # Update calculated diameter and length from file parameters
        print("DEBUG: Starting parameter updates...")

        # Try to get inputs through their groups
        options_group = inputs.itemById('options_group')
        if options_group:
            print("DEBUG: Found options group")
            options_inputs = options_group.children
            diameter_input = adsk.core.ValueCommandInput.cast(
                options_inputs.itemById('diameter'))
            length_input = adsk.core.ValueCommandInput.cast(
                options_inputs.itemById('length'))
        else:
            print("DEBUG: Options group not found, trying direct access")
            diameter_input = adsk.core.ValueCommandInput.cast(
                inputs.itemById('diameter'))
            length_input = adsk.core.ValueCommandInput.cast(
                inputs.itemById('length'))

        print(f"DEBUG: diameter_input = {diameter_input}")
        print(f"DEBUG: length_input = {length_input}")

        # Debug: Check if inputs exist at all
        print(
            f"DEBUG: Checking if 'diameter' input exists: {inputs.itemById('diameter') is not None}")
        print(
            f"DEBUG: Checking if 'length' input exists: {inputs.itemById('length') is not None}")

        # Debug: List all input IDs to see what's available
        print("DEBUG: All available input IDs:")
        for i in range(inputs.count):
            input_item = inputs.item(i)
            print(f"  - {input_item.id}: {type(input_item)}")

        if diameter_input and data:
            print("DEBUG: Checking diameter parameter...")
            if 'diameter_mm' in parameters:
                print("DEBUG: Found diameter in parameters")
                # Use diameter from file parameters
                diameter_value = parameters['diameter_mm']
                print(
                    f"DEBUG: Setting diameter from file: {diameter_value} mm")
                # Convert mm to cm since Fusion treats input as cm
                diameter_cm = diameter_value / 10
                print(f"DEBUG: Converted to: {diameter_cm} cm")

                # Try multiple approaches to set the value
                try:
                    diameter_input.value = diameter_cm
                    print(f"DEBUG: Set diameter_input.value = {diameter_cm}")
                except Exception as e:
                    print(f"DEBUG: Failed to set value directly: {e}")

                try:
                    diameter_input.expression = f"{diameter_cm}"
                    print(
                        f"DEBUG: Set diameter_input.expression = {diameter_cm}")
                except Exception as e:
                    print(f"DEBUG: Failed to set expression: {e}")

                print(
                    f"DEBUG: Final diameter_input.value = {diameter_input.value}")
                print(
                    f"DEBUG: Final diameter_input.expression = {diameter_input.expression}")
            else:
                print("DEBUG: No diameter_mm found in parameters")
                # Calculate diameter from wave width
                wave_widths = [float(row.get('wave_width_mm', 0))
                               for row in data]
                avg_wave_width = sum(wave_widths) / \
                    len(wave_widths) if wave_widths else 0
                cols_in_first_ring = len(
                    [row for row in data if row['ring'] == rings[0]])
                calculated_diameter = (
                    avg_wave_width * cols_in_first_ring) / math.pi
                diameter_cm = calculated_diameter / 10
                diameter_input.expression = f"{diameter_cm} cm"
                print(
                    f"DEBUG: Calculated diameter set to: {calculated_diameter} mm = {diameter_cm} cm")

        if length_input and 'length_mm' in parameters:
            print("DEBUG: Updating length parameter...")
            # Convert mm to cm since Fusion treats input as cm
            length_value = parameters['length_mm']
            print(f"DEBUG: Setting length from file: {length_value} mm")
            length_cm = length_value / 10
            length_input.expression = f"{length_cm} cm"
            print(
                f"DEBUG: Length input expression set to: {length_input.expression}")
            print(f"DEBUG: Length input value is now: {length_input.value}")
        else:
            print(
                f"DEBUG: Length update skipped. length_input={length_input}, 'length_mm' in parameters={('length_mm' in parameters)}")

        # Update additional parameters from file
        num_rings_input = adsk.core.IntegerSpinnerCommandInput.cast(
            inputs.itemById('num_rings'))
        crowns_per_ring_input = adsk.core.IntegerSpinnerCommandInput.cast(
            inputs.itemById('crowns_per_ring'))
        angle_input = adsk.core.ValueCommandInput.cast(
            inputs.itemById('angle_per_crown'))

        if num_rings_input and 'num_rings' in parameters:
            num_rings_input.value = int(parameters['num_rings'])

        if crowns_per_ring_input and 'crowns_per_ring' in parameters:
            crowns_per_ring_input.value = int(parameters['crowns_per_ring'])

            if angle_input and 'angle_per_crown_deg' in parameters:
                # Convert degrees to radians for Fusion 360
                angle_deg = parameters['angle_per_crown_deg']
                angle_rad = math.radians(angle_deg)
                angle_input.value = angle_rad
                print(f"DEBUG: Angle set to: {angle_deg}° = {angle_rad} rad")

        # Force dialog refresh to ensure UI updates
        print("DEBUG: Forcing dialog refresh...")
        try:
            # Force UI refresh by triggering a command update
            if hasattr(inputs.command, 'doExecutePreview'):
                inputs.command.doExecutePreview()
            else:
                # Alternative refresh method
                if hasattr(inputs, 'command'):
                    inputs.command.executePreview()
        except Exception as refresh_error:
            print(
                f"DEBUG: Dialog refresh error (non-critical): {refresh_error}")

    except Exception as e:
        if status_text:
            status_text.text = f'Error: {str(e)}'
        futil.log(f'Error in update_data_preview: {traceback.format_exc()}')


def read_excel_data(file_path):
    """Read Excel data from WaveInputsByColumn sheet"""
    try:
        # For now, provide a simple CSV alternative or require manual data entry
        # This is a placeholder that would work with CSV data

        # Check if it's a CSV or JSON file instead
        if file_path.lower().endswith('.csv'):
            csv_data = read_csv_data(file_path)
            return {'data': csv_data, 'parameters': {}}
        elif file_path.lower().endswith('.json'):
            return read_json_data(file_path)

        # For Excel files, we'll need openpyxl or pandas
        # Try to import openpyxl
        try:
            import openpyxl  # type: ignore

            wb = openpyxl.load_workbook(file_path, data_only=True)

            if 'WaveInputsByColumn' not in wb.sheetnames:
                raise Exception(
                    "Sheet 'WaveInputsByColumn' not found in Excel file.")

            ws = wb['WaveInputsByColumn']

            # Get header row
            headers = []
            for cell in ws[1]:
                if cell.value:
                    headers.append(str(cell.value).strip())

            # Read data rows
            data = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] is not None:  # Skip empty rows
                    row_data = {}
                    for i, header in enumerate(headers):
                        if i < len(row) and row[i] is not None:
                            value = row[i]
                            try:
                                # Convert to appropriate type
                                if header in ['ring', 'col']:
                                    row_data[header] = int(float(str(value)))
                                else:
                                    row_data[header] = float(str(value))
                            except (ValueError, TypeError):
                                row_data[header] = str(value)
                        else:
                            row_data[header] = 0.0 if header not in [
                                'ring', 'col'] else 0
                    data.append(row_data)

            wb.close()
            return {'data': data, 'parameters': {}}

        except ImportError:
            # If openpyxl is not available, suggest alternative
            raise Exception(
                "openpyxl library not found. Please install openpyxl or save your Excel file as CSV format.")

    except Exception as e:
        futil.log(f'Error reading Excel file: {traceback.format_exc()}')
        raise


def read_csv_data(file_path):
    """Read CSV data as alternative to Excel"""
    try:
        import csv
        data = []

        with open(file_path, 'r', newline='', encoding='utf-8') as csvfile:
            reader = csv.DictReader(csvfile)

            for row in reader:
                row_data = {}
                for key, value in row.items():
                    try:
                        if key in ['ring', 'col']:
                            row_data[key] = int(float(value))
                        else:
                            row_data[key] = float(value)
                    except (ValueError, TypeError):
                        row_data[key] = 0.0 if key not in [
                            'ring', 'col'] else 0
                data.append(row_data)

        return data

    except Exception as e:
        futil.log(f'Error reading CSV file: {traceback.format_exc()}')
        raise


def read_json_data(file_path):
    """Read JSON data with stent frame structure"""
    try:
        import json

        with open(file_path, 'r', encoding='utf-8') as jsonfile:
            json_data = json.load(jsonfile)

        # Extract parameters for diameter and length
        parameters = {}
        if 'parameters' in json_data:
            parameters = json_data['parameters']

        # Check for new format with 'cells' data
        if 'cells' in json_data:
            data = json_data['cells']
            print("DEBUG: Found new JSON format with 'cells' data")

            # Convert to the expected format if needed
            formatted_data = []
            for row in data:
                # Ensure all numeric values are properly typed
                row_data = {}
                for key, value in row.items():
                    if value is None:
                        # Handle null values
                        if key in ['ring', 'col']:
                            row_data[key] = 0
                        else:
                            row_data[key] = 0.0
                    elif key in ['ring', 'col', 'linked_above', 'linked_below']:
                        try:
                            row_data[key] = int(
                                value) if value is not None else 0
                        except (ValueError, TypeError):
                            row_data[key] = 0
                    elif key in ['chord_top_centerline', 'chord_bottom_centerline', 'chord_top_outer', 'chord_bottom_outer']:
                        # Keep chord coordinate arrays as-is
                        row_data[key] = value
                    else:
                        try:
                            row_data[key] = float(
                                value) if value is not None else 0.0
                        except (ValueError, TypeError):
                            row_data[key] = 0.0
                formatted_data.append(row_data)

            # Return both data and parameters
            return {
                'data': formatted_data,
                'parameters': parameters
            }
        # Extract wave_inputs_by_column data (old format)
        elif 'wave_inputs_by_column' in json_data:
            data = json_data['wave_inputs_by_column']
            print("DEBUG: Found old JSON format with 'wave_inputs_by_column' data")

            # Convert to the expected format if needed
            formatted_data = []
            for row in data:
                # Ensure all numeric values are properly typed
                row_data = {}
                for key, value in row.items():
                    if value is None:
                        # Handle null values
                        if key in ['ring', 'col']:
                            row_data[key] = 0
                        else:
                            row_data[key] = 0.0
                    elif key in ['ring', 'col', 'linked_above', 'linked_below']:
                        try:
                            row_data[key] = int(
                                value) if value is not None else 0
                        except (ValueError, TypeError):
                            row_data[key] = 0
                    else:
                        try:
                            row_data[key] = float(
                                value) if value is not None else 0.0
                        except (ValueError, TypeError):
                            row_data[key] = 0.0
                formatted_data.append(row_data)

            # Return both data and parameters
            return {
                'data': formatted_data,
                'parameters': parameters
            }
        else:
            raise Exception(
                "No 'cells' or 'wave_inputs_by_column' data found in JSON file")

    except Exception as e:
        futil.log(f'Error reading JSON file: {traceback.format_exc()}')
        raise


def process_excel_file(file_path, diameter_mm, length_mm=None, draw_construction=True, draw_chords=True, create_points=False):
    """Process the Excel file and create the stent frame sketch"""
    try:
        # Read Excel data (now returns dict with 'data' and 'parameters')
        file_data = read_excel_data(file_path)
        data = file_data['data']
        parameters = file_data.get('parameters', {})

        if not data:
            raise Exception("No data found in Excel file")

        # Use diameter from JSON parameters if available, otherwise use provided value
        if 'diameter_mm' in parameters:
            diameter_mm = parameters['diameter_mm']
            print(f"Using diameter from file: {diameter_mm} mm")

        # Use length from JSON parameters if available
        length_mm = parameters.get('length_mm', None)
        if length_mm:
            print(f"Using length from file: {length_mm} mm")

        # Get Fusion objects
        app = adsk.core.Application.get()
        ui = app.userInterface
        design = adsk.fusion.Design.cast(app.activeProduct)

        if not design:
            raise Exception(
                "Please create a new design or open an existing one before running this command.")

        root = design.rootComponent

        # Create new sketch
        sketch = root.sketches.add(root.xYConstructionPlane)

        # Analyze data structure
        rings = sorted(set(row['ring'] for row in data))

        # Calculate stent dimensions
        first_ring_data = [row for row in data if row['ring'] == rings[0]]
        cols_per_ring = len(first_ring_data)

        # Calculate total length and ring positions from absolute Y positions
        total_length_mm = 0
        ring_positions = {}

        # Check if we have absolute Y position data
        has_absolute_positions = any(
            'y_top_border_mm' in row and 'y_bottom_border_mm' in row for row in data)

        if has_absolute_positions:
            print("DEBUG: Using absolute Y positions from data")
            # Use absolute Y positions from the data
            for ring_num in rings:
                ring_data = [row for row in data if row['ring'] == ring_num]

                # Get min top and max bottom for this ring
                top_positions = [row['y_top_border_mm']
                                 for row in ring_data if 'y_top_border_mm' in row]
                bottom_positions = [row['y_bottom_border_mm']
                                    for row in ring_data if 'y_bottom_border_mm' in row]

                if top_positions and bottom_positions:
                    start_y = min(top_positions)
                    end_y = max(bottom_positions)
                    ring_height = end_y - start_y

                    ring_positions[ring_num] = {
                        'center_y': start_y + ring_height / 2,
                        'start_y': start_y,
                        'end_y': end_y,
                        'height': ring_height
                    }

                    # Update total length
                    total_length_mm = max(total_length_mm, end_y)
        else:
            print("DEBUG: Calculating Y positions from wave heights and gaps")
            # Fallback to calculated positions
            current_y = 0
            for ring_num in rings:
                ring_data = [row for row in data if row['ring'] == ring_num]
                ring_height = max(row['wave_height_mm'] for row in ring_data)

                ring_positions[ring_num] = {
                    'center_y': current_y + ring_height / 2,
                    'start_y': current_y,
                    'end_y': current_y + ring_height,
                    'height': ring_height
                }

                current_y += ring_height

                # Add gap after ring (except for last ring)
                if ring_num < max(rings):
                    # Use gap_below from current ring data
                    gap_below = max(row.get('gap_below_mm', 0)
                                    for row in ring_data)
                    current_y += gap_below

            total_length_mm = current_y

        # Use provided length if available, otherwise use calculated length
        if length_mm is not None:
            print(
                f"DEBUG: Using provided length: {length_mm} mm (calculated was: {total_length_mm:.3f} mm)")
            total_length_mm = length_mm
        else:
            print(f"DEBUG: Using calculated length: {total_length_mm:.3f} mm")

        # Calculate width from diameter
        width_mm = diameter_mm * math.pi

        # Set sketch name
        sketch.name = f'Stent Frame from Excel - {len(rings)} rings, {cols_per_ring} cols'

        lines = sketch.sketchCurves.sketchLines

        # Convert mm to cm for Fusion API
        def mm_to_cm(x):
            return x * 0.1

        # Draw border
        if draw_construction:
            # Left border
            lines.addByTwoPoints(
                adsk.core.Point3D.create(0, 0, 0),
                adsk.core.Point3D.create(0, mm_to_cm(total_length_mm), 0)
            ).isConstruction = True

            # Right border
            lines.addByTwoPoints(
                adsk.core.Point3D.create(mm_to_cm(width_mm), 0, 0),
                adsk.core.Point3D.create(
                    mm_to_cm(width_mm), mm_to_cm(total_length_mm), 0)
            ).isConstruction = True

            # Top border
            lines.addByTwoPoints(
                adsk.core.Point3D.create(0, mm_to_cm(total_length_mm), 0),
                adsk.core.Point3D.create(
                    mm_to_cm(width_mm), mm_to_cm(total_length_mm), 0)
            ).isConstruction = True

            # Bottom border
            lines.addByTwoPoints(
                adsk.core.Point3D.create(0, 0, 0),
                adsk.core.Point3D.create(mm_to_cm(width_mm), 0, 0)
            ).isConstruction = True

        # Draw ring boundaries
        if draw_construction:
            for ring_num, ring_info in ring_positions.items():
                # Ring start line
                lines.addByTwoPoints(
                    adsk.core.Point3D.create(
                        0, mm_to_cm(ring_info['start_y']), 0),
                    adsk.core.Point3D.create(
                        mm_to_cm(width_mm), mm_to_cm(ring_info['start_y']), 0)
                ).isConstruction = True

                # Ring end line
                lines.addByTwoPoints(
                    adsk.core.Point3D.create(
                        0, mm_to_cm(ring_info['end_y']), 0),
                    adsk.core.Point3D.create(
                        mm_to_cm(width_mm), mm_to_cm(ring_info['end_y']), 0)
                ).isConstruction = True

        # Draw column boundaries
        if draw_construction:
            col_spacing = width_mm / cols_per_ring
            for col in range(1, cols_per_ring):
                x_pos = col * col_spacing
                lines.addByTwoPoints(
                    adsk.core.Point3D.create(mm_to_cm(x_pos), 0, 0),
                    adsk.core.Point3D.create(
                        mm_to_cm(x_pos), mm_to_cm(total_length_mm), 0)
                ).isConstruction = True

        # Draw chord lines based on Excel data
        if draw_chords:
            col_spacing = width_mm / cols_per_ring

            for row in data:
                ring_num = row['ring']
                col_num = row['col']

                # Check if this row has crown chord coordinate data (new format)
                if 'chord_top_centerline' in row and 'chord_bottom_centerline' in row:
                    print(
                        f"DEBUG: Drawing crown chords using coordinates for ring {ring_num}, col {col_num}")

                    # Draw top crown chord using centerline coordinates
                    chord_top = row.get('chord_top_centerline', [])
                    if len(chord_top) >= 2:
                        start_point = chord_top[0]  # [x, y]
                        end_point = chord_top[1]    # [x, y]

                        lines.addByTwoPoints(
                            adsk.core.Point3D.create(
                                mm_to_cm(start_point[0]), mm_to_cm(start_point[1]), 0),
                            adsk.core.Point3D.create(
                                mm_to_cm(end_point[0]), mm_to_cm(end_point[1]), 0)
                        ).isConstruction = True

                    # Draw bottom crown chord using centerline coordinates
                    chord_bottom = row.get('chord_bottom_centerline', [])
                    if len(chord_bottom) >= 2:
                        start_point = chord_bottom[0]  # [x, y]
                        end_point = chord_bottom[1]    # [x, y]

                        lines.addByTwoPoints(
                            adsk.core.Point3D.create(
                                mm_to_cm(start_point[0]), mm_to_cm(start_point[1]), 0),
                            adsk.core.Point3D.create(
                                mm_to_cm(end_point[0]), mm_to_cm(end_point[1]), 0)
                        ).isConstruction = True

                    # Optionally draw outer edge chords for keep-out zones
                    if 'chord_top_outer' in row and 'chord_bottom_outer' in row:
                        # Draw top crown chord outer edge
                        chord_top_outer = row.get('chord_top_outer', [])
                        if len(chord_top_outer) >= 2:
                            start_point = chord_top_outer[0]  # [x, y]
                            end_point = chord_top_outer[1]    # [x, y]

                            outer_line = lines.addByTwoPoints(
                                adsk.core.Point3D.create(
                                    mm_to_cm(start_point[0]), mm_to_cm(start_point[1]), 0),
                                adsk.core.Point3D.create(
                                    mm_to_cm(end_point[0]), mm_to_cm(end_point[1]), 0)
                            )
                            outer_line.isConstruction = True
                            # Make outer edge lines a different style if possible

                        # Draw bottom crown chord outer edge
                        chord_bottom_outer = row.get('chord_bottom_outer', [])
                        if len(chord_bottom_outer) >= 2:
                            start_point = chord_bottom_outer[0]  # [x, y]
                            end_point = chord_bottom_outer[1]    # [x, y]

                            outer_line = lines.addByTwoPoints(
                                adsk.core.Point3D.create(
                                    mm_to_cm(start_point[0]), mm_to_cm(start_point[1]), 0),
                                adsk.core.Point3D.create(
                                    mm_to_cm(end_point[0]), mm_to_cm(end_point[1]), 0)
                            )
                            outer_line.isConstruction = True

                # Fallback to old method if no crown chord coordinates available
                elif ring_num in ring_positions:
                    ring_info = ring_positions[ring_num]

                    # Calculate column center position
                    col_center_x = (col_num + 0.5) * col_spacing

                    # Get chord and sagitta data
                    upper_chord = row.get('upper_chord_center_mm', 0)
                    upper_sagitta = row.get('upper_sagitta_center_mm', 0)
                    lower_chord = row.get('lower_chord_center_mm', 0)
                    lower_sagitta = row.get('lower_sagitta_center_mm', 0)

                    # Draw upper chord line
                    if upper_chord > 0:
                        chord_half_length = upper_chord / 2
                        chord_y = ring_info['center_y'] + \
                            ring_info['height']/2 - upper_sagitta

                        lines.addByTwoPoints(
                            adsk.core.Point3D.create(
                                mm_to_cm(col_center_x - chord_half_length), mm_to_cm(chord_y), 0),
                            adsk.core.Point3D.create(
                                mm_to_cm(col_center_x + chord_half_length), mm_to_cm(chord_y), 0)
                        ).isConstruction = True

                    # Draw lower chord line
                    if lower_chord > 0:
                        chord_half_length = lower_chord / 2
                        chord_y = ring_info['center_y'] - \
                            ring_info['height']/2 + lower_sagitta

                        lines.addByTwoPoints(
                            adsk.core.Point3D.create(
                                mm_to_cm(col_center_x - chord_half_length), mm_to_cm(chord_y), 0),
                            adsk.core.Point3D.create(
                                mm_to_cm(col_center_x + chord_half_length), mm_to_cm(chord_y), 0)
                        ).isConstruction = True

        # Draw individual cell frames using absolute Y positions
        if has_absolute_positions and draw_construction:
            print("DEBUG: Drawing individual cell frames using absolute Y positions")
            col_spacing = width_mm / cols_per_ring

            for row in data:
                ring_num = row['ring']
                col_num = row['col']

                # Check if this row has absolute position data
                if 'y_top_border_mm' in row and 'y_bottom_border_mm' in row:
                    y_top = row['y_top_border_mm']
                    y_bottom = row['y_bottom_border_mm']

                    # Calculate column boundaries
                    x_left = col_num * col_spacing
                    x_right = (col_num + 1) * col_spacing

                    # Draw cell frame rectangle (construction lines)
                    # Top border of cell
                    lines.addByTwoPoints(
                        adsk.core.Point3D.create(
                            mm_to_cm(x_left), mm_to_cm(y_top), 0),
                        adsk.core.Point3D.create(
                            mm_to_cm(x_right), mm_to_cm(y_top), 0)
                    ).isConstruction = True

                    # Bottom border of cell
                    lines.addByTwoPoints(
                        adsk.core.Point3D.create(
                            mm_to_cm(x_left), mm_to_cm(y_bottom), 0),
                        adsk.core.Point3D.create(
                            mm_to_cm(x_right), mm_to_cm(y_bottom), 0)
                    ).isConstruction = True

                    # Left border of cell
                    lines.addByTwoPoints(
                        adsk.core.Point3D.create(
                            mm_to_cm(x_left), mm_to_cm(y_top), 0),
                        adsk.core.Point3D.create(
                            mm_to_cm(x_left), mm_to_cm(y_bottom), 0)
                    ).isConstruction = True

                    # Right border of cell
                    lines.addByTwoPoints(
                        adsk.core.Point3D.create(
                            mm_to_cm(x_right), mm_to_cm(y_top), 0),
                        adsk.core.Point3D.create(
                            mm_to_cm(x_right), mm_to_cm(y_bottom), 0)
                    ).isConstruction = True

        # Create points at intersections if requested
        if create_points:
            col_spacing = width_mm / cols_per_ring

            for ring_num, ring_info in ring_positions.items():
                for col in range(cols_per_ring + 1):
                    x_pos = col * col_spacing

                    # Points at ring boundaries
                    sketch.sketchPoints.add(adsk.core.Point3D.create(
                        mm_to_cm(x_pos), mm_to_cm(ring_info['start_y']), 0))
                    sketch.sketchPoints.add(adsk.core.Point3D.create(
                        mm_to_cm(x_pos), mm_to_cm(ring_info['end_y']), 0))

        # Show summary
        rings_count = len(rings)
        data_points = len(data)
        positioning_method = "Absolute Y positions" if has_absolute_positions else "Calculated positions"
        cell_frames_drawn = has_absolute_positions and draw_construction

        # Check if crown chord coordinates are available
        has_crown_chords = any('chord_top_centerline' in row for row in data)
        chord_method = "Crown chord coordinates" if has_crown_chords else "Calculated sagitta positions"

        ui.messageBox(
            f'Stent frame created from Excel data!\n\n'
            f'• File: {os.path.basename(file_path)}\n'
            f'• Data points processed: {data_points}\n'
            f'• Rings: {rings_count} (rings {min(rings)} to {max(rings)})\n'
            f'• Columns per ring: {cols_per_ring}\n'
            f'• Calculated diameter: {diameter_mm:.2f} mm\n'
            f'• Calculated length: {total_length_mm:.2f} mm\n'
            f'• Width (circumference): {width_mm:.2f} mm\n'
            f'• Positioning method: {positioning_method}\n'
            f'• Construction lines: {"Yes" if draw_construction else "No"}\n'
            f'• Chord lines: {"Yes" if draw_chords else "No"}\n'
            f'• Chord method: {chord_method}\n'
            f'• Individual cell frames: {"Yes" if cell_frames_drawn else "No"}\n'
            f'• Sketch points: {"Yes" if create_points else "No"}'
        )

    except Exception as e:
        futil.log(f'Error in process_excel_file: {traceback.format_exc()}')
        raise

# This event handler is called when the user interacts with any of the inputs in the dialog
# which allows you to verify that all of the inputs are valid and enables the OK button.


def command_validate_input(args: adsk.core.ValidateInputsEventArgs):
    # General logging for debug.
    futil.log(f'{CMD_NAME} Validate Input Event')

    inputs = args.inputs

    # Apply stored parameters if available
    global global_parameters
    if global_parameters:
        print("DEBUG: Attempting to apply stored parameters in validate event")

        # Try to get diameter and length inputs
        diameter_input = adsk.core.ValueCommandInput.cast(
            inputs.itemById('diameter'))
        length_input = adsk.core.ValueCommandInput.cast(
            inputs.itemById('length'))

        print(f"DEBUG: In validate - diameter_input = {diameter_input}")
        print(f"DEBUG: In validate - length_input = {length_input}")

        if diameter_input and 'diameter_mm' in global_parameters:
            try:
                # Convert mm to cm since Fusion treats input as cm
                diameter_cm = global_parameters['diameter_mm'] / 10
                diameter_input.value = diameter_cm
                print(
                    f"DEBUG: Applied diameter: {global_parameters['diameter_mm']} mm = {diameter_cm} cm")
            except Exception as e:
                print(f"DEBUG: Error setting diameter: {e}")

        if length_input and 'length_mm' in global_parameters:
            try:
                # Convert mm to cm since Fusion treats input as cm
                length_cm = global_parameters['length_mm'] / 10
                length_input.value = length_cm
                print(
                    f"DEBUG: Applied length: {global_parameters['length_mm']} mm = {length_cm} cm")
            except Exception as e:
                print(f"DEBUG: Error setting length: {e}")

        # Apply other parameters
        num_rings_input = adsk.core.IntegerSpinnerCommandInput.cast(
            inputs.itemById('num_rings'))
        if num_rings_input and 'num_rings' in global_parameters:
            try:
                num_rings_input.value = int(global_parameters['num_rings'])
                print(
                    f"DEBUG: Applied num_rings: {global_parameters['num_rings']}")
            except Exception as e:
                print(f"DEBUG: Error setting num_rings: {e}")

        crowns_per_ring_input = adsk.core.IntegerSpinnerCommandInput.cast(
            inputs.itemById('crowns_per_ring'))
        if crowns_per_ring_input and 'crowns_per_ring' in global_parameters:
            try:
                crowns_per_ring_input.value = int(
                    global_parameters['crowns_per_ring'])
                print(
                    f"DEBUG: Applied crowns_per_ring: {global_parameters['crowns_per_ring']}")
            except Exception as e:
                print(f"DEBUG: Error setting crowns_per_ring: {e}")

        # Clear the parameters after applying them once
        global_parameters = {}

    # Check if file path is provided and exists
    file_path_input = adsk.core.StringValueCommandInput.cast(
        inputs.itemById('file_path'))

    if file_path_input and file_path_input.value and os.path.exists(file_path_input.value):
        args.areInputsValid = True
    else:
        args.areInputsValid = False

# This event handler is called when the command terminates.


def command_destroy(args: adsk.core.CommandEventArgs):
    # General logging for debug.
    futil.log(f'{CMD_NAME} Command Destroy Event')

    global local_handlers
    local_handlers = []
