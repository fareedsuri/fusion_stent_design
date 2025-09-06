import adsk.core
import adsk.fusion
import adsk.cam
import traceback
import os
import sys
import math

# Import fusion utilities
from ...lib import fusionAddInUtils as futil
from ... import config

# Add the parent directory to sys.path to import crown_arc module
parent_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
if parent_dir not in sys.path:
    sys.path.append(parent_dir)

# Import the crown_arc module for calculations
try:
    import crown_arc
except ImportError:
    crown_arc = None

# TODO *** Define the location of the command ***
# This is done by declaring the space, the tab, and the panel.
CMD_ID = f'{config.COMPANY_NAME}_{config.ADDIN_NAME}_excelProcessor'
CMD_NAME = 'Process Excel Stent Data'
CMD_Description = 'Load and process Excel file with detailed stent frame data'

# Specify that the command will be promoted to the panel.
IS_PROMOTED = True

# TODO *** Define the location of the command ***
# This is done by declaring the space, the tab, and the panel.
WORKSPACE_ID = 'FusionSolidEnvironment'
PANEL_ID = 'SolidCreatePanel'
COMMAND_BESIDE_ID = 'ScriptsManagerCommand'

# Resource file folders relative to this file.
ICON_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'resources', '')

# Local list of event handlers used to maintain a reference so
# they are not released and garbage collected.
local_handlers = []

# Global variables
excel_file_path = ""

def start():
    # Create a command Definition.
    cmd_def = adsk.core.Application.get().userInterface.commandDefinitions.addButtonDefinition(CMD_ID, CMD_NAME, CMD_Description, ICON_FOLDER)

    # Define an event handler for the command created event. It will be called when the button is clicked.
    futil.add_handler(cmd_def.commandCreated, command_created)

    # ******** Add a button into the UI so the user can run the command. ********
    # Get the target workspace the button will be created in.
    workspace = adsk.core.Application.get().userInterface.workspaces.itemById(WORKSPACE_ID)

    # Get the panel the button will be created in.
    panel = workspace.toolbarPanels.itemById(PANEL_ID)

    # Create the button command control in the UI after the specified existing command.
    control = panel.controls.addCommand(cmd_def, COMMAND_BESIDE_ID, False)

    # Specify if the command is promoted to the main toolbar.
    control.isPromoted = IS_PROMOTED

def stop():
    # Get the various UI elements for this command
    workspace = adsk.core.Application.get().userInterface.workspaces.itemById(WORKSPACE_ID)
    panel = workspace.toolbarPanels.itemById(PANEL_ID)
    command_control = panel.controls.itemById(CMD_ID)
    command_definition = adsk.core.Application.get().userInterface.commandDefinitions.itemById(CMD_ID)

    # Delete the button command control
    if command_control:
        command_control.deleteMe()

    # Delete the command definition
    if command_definition:
        command_definition.deleteMe()

# Function that is called when a user clicks the corresponding button in the UI.
# This defines the contents of the command dialog and connects to the command related events.
def command_created(args: adsk.core.CommandCreatedEventArgs):
    # General logging for debug.
    futil.log(f'{CMD_NAME} Command Created Event')

    # https://help.autodesk.com/view/fusion360/ENU/?contextId=CommandInputs
    inputs = args.command.commandInputs
    
    # Create a default value using a string. This method also accepts a real as a parameter.
    # If a real is provided, the value will be assumed to be in current user units.
    
    # Add a message to describe the command
    message_input = inputs.addTextBoxCommandInput('message', '', 
        'This command processes Excel files with detailed stent frame data.\n\n'
        'The Excel file should contain a sheet named "WaveInputsByColumn" with\n'
        'columns: ring, col, wave_height_mm, wave_width_mm, strut_width_mm,\n'
        'gap_above_mm, gap_below_mm, upper_chord_center_mm, upper_sagitta_center_mm,\n'
        'upper_outer_sagitta_mm, lower_chord_center_mm, lower_sagitta_center_mm,\n'
        'lower_outer_sagitta_mm, theta_deg, Rc_mm', 6, True)
    
    # File selection group
    file_group = inputs.addGroupCommandInput('file_group', 'Excel File Selection')
    file_group.isExpanded = True
    file_group_inputs = file_group.children
    
    # Add a text input for the file path
    file_path_input = file_group_inputs.addStringValueInput('file_path', 'Excel File Path', '')
    file_path_input.tooltip = 'Path to the Excel file containing stent data'
    
    # Add a button to browse for file
    browse_button = file_group_inputs.addBoolValueInput('browse_file', 'Browse for File', False, '', False)
    browse_button.tooltip = 'Click to browse for Excel file'
    
    # Processing options group
    options_group = inputs.addGroupCommandInput('options_group', 'Processing Options')
    options_group.isExpanded = True
    options_group_inputs = options_group.children
    
    # Add diameter input (will be calculated from data)
    diameter_input = options_group_inputs.addValueInput('diameter', 'Stent Diameter', 'mm', 
                                                      adsk.core.ValueInput.createByReal(5.0))
    diameter_input.tooltip = 'Stent diameter in millimeters (will be calculated from wave data)'
    
    # Add option to draw construction lines
    draw_construction = options_group_inputs.addBoolValueInput('draw_construction', 'Draw Construction Lines', True, '', True)
    draw_construction.tooltip = 'Draw construction lines showing wave boundaries and centers'
    
    # Add option to draw chord lines
    draw_chords = options_group_inputs.addBoolValueInput('draw_chords', 'Draw Chord Lines', True, '', True)
    draw_chords.tooltip = 'Draw chord lines at sagitta positions for each crown'
    
    # Add option to create sketch points
    create_points = options_group_inputs.addBoolValueInput('create_points', 'Create Sketch Points', True, '', False)
    create_points.tooltip = 'Create sketch points at key intersections'
    
    # Data preview group
    preview_group = inputs.addGroupCommandInput('preview_group', 'Data Preview')
    preview_group.isExpanded = False
    preview_group_inputs = preview_group.children
    
    # Add a text box for data preview
    preview_text = preview_group_inputs.addTextBoxCommandInput('data_preview', '', 
        'No file selected. Use "Browse for File" to select an Excel file.', 4, True)
    
    # Status group
    status_group = inputs.addGroupCommandInput('status_group', 'Status')
    status_group.isExpanded = False
    status_group_inputs = status_group.children
    
    # Add status text
    status_text = status_group_inputs.addTextBoxCommandInput('status', '', 'Ready to process Excel file.', 2, True)

    # Connect to the events that are needed by this command.
    futil.add_handler(args.command.execute, command_execute, local_handlers=local_handlers)
    futil.add_handler(args.command.inputChanged, command_input_changed, local_handlers=local_handlers)
    futil.add_handler(args.command.executePreview, command_preview, local_handlers=local_handlers)
    futil.add_handler(args.command.validateInputs, command_validate_input, local_handlers=local_handlers)
    futil.add_handler(args.command.destroy, command_destroy, local_handlers=local_handlers)

# This event handler is called when the user clicks the OK button in the command dialog or 
# is immediately called after the created event not command inputs were created for the dialog.
def command_execute(args: adsk.core.CommandEventArgs):
    # General logging for debug.
    futil.log(f'{CMD_NAME} Command Execute Event')

    # TODO *** Do something useful here in the command. ***
    
    # Get the values from the command inputs.
    inputs = args.command.commandInputs
    
    file_path_input = adsk.core.StringValueCommandInput.cast(inputs.itemById('file_path'))
    diameter_input = adsk.core.ValueCommandInput.cast(inputs.itemById('diameter'))
    draw_construction_input = adsk.core.BoolValueCommandInput.cast(inputs.itemById('draw_construction'))
    draw_chords_input = adsk.core.BoolValueCommandInput.cast(inputs.itemById('draw_chords'))
    create_points_input = adsk.core.BoolValueCommandInput.cast(inputs.itemById('create_points'))
    
    if not file_path_input.value:
        adsk.core.Application.get().userInterface.messageBox('Please select an Excel file first.')
        return
        
    # Process the Excel file and create the stent frame
    try:
        process_excel_file(
            file_path=file_path_input.value,
            diameter_mm=diameter_input.value * 10,  # Convert cm to mm
            draw_construction=draw_construction_input.value,
            draw_chords=draw_chords_input.value,
            create_points=create_points_input.value
        )
    except Exception as e:
        adsk.core.Application.get().userInterface.messageBox(f'Error processing Excel file: {str(e)}')
        futil.log(f'Error in command_execute: {traceback.format_exc()}')

# This event handler is called when the command needs to compute a new preview in the graphics window.
def command_preview(args: adsk.core.CommandEventArgs):
    # General logging for debug.
    futil.log(f'{CMD_NAME} Command Preview Event')
    inputs = args.command.commandInputs

# This event handler is called when the user changes anything in the command dialog
# allowing you to modify values of other inputs based on that change.
def command_input_changed(args: adsk.core.InputChangedEventArgs):
    changed_input = args.input
    inputs = args.inputs
    
    futil.log(f'{CMD_NAME} Input Changed Event fired from a change to {changed_input.id}')
    
    # Handle file browse button
    if changed_input.id == 'browse_file':
        browse_button = adsk.core.BoolValueCommandInput.cast(changed_input)
        if browse_button.value:
            # Reset the button
            browse_button.value = False
            
            # Open file dialog
            file_dialog = adsk.core.Application.get().userInterface.createFileDialog()
            file_dialog.isMultiSelectEnabled = False
            file_dialog.title = "Select Excel File"
            file_dialog.filter = "Excel files (*.xlsx);;All files (*.*)"
            
            if file_dialog.showOpen() == adsk.core.DialogResults.DialogOK:
                file_path = file_dialog.filename
                file_path_input = adsk.core.StringValueCommandInput.cast(inputs.itemById('file_path'))
                file_path_input.value = file_path
                
                # Update preview
                update_data_preview(inputs, file_path)
    
    # Handle file path changes
    elif changed_input.id == 'file_path':
        file_path_input = adsk.core.StringValueCommandInput.cast(changed_input)
        if file_path_input.value and os.path.exists(file_path_input.value):
            update_data_preview(inputs, file_path_input.value)

def update_data_preview(inputs, file_path):
    """Update the data preview text box with Excel file information"""
    preview_text = None
    status_text = None
    
    try:
        preview_text = adsk.core.TextBoxCommandInput.cast(inputs.itemById('data_preview'))
        status_text = adsk.core.TextBoxCommandInput.cast(inputs.itemById('status'))
        
        if not os.path.exists(file_path):
            if preview_text:
                preview_text.text = 'File not found.'
            if status_text:
                status_text.text = 'Error: File not found.'
            return
            
        # Try to read Excel file
        data = read_excel_data(file_path)
        
        if data is None or len(data) == 0:
            if preview_text:
                preview_text.text = 'No data found in WaveInputsByColumn sheet.'
            if status_text:
                status_text.text = 'Error: No valid data found.'
            return
            
        # Create preview text
        rings = sorted(set(row['ring'] for row in data))
        cols_per_ring = {}
        for ring in rings:
            ring_data = [row for row in data if row['ring'] == ring]
            cols_per_ring[ring] = len(ring_data)
            
        total_rows = len(data)
        
        preview = f"Excel file loaded successfully!\n\n"
        preview += f"Total data rows: {total_rows}\n"
        preview += f"Rings found: {len(rings)} (rings {min(rings)} to {max(rings)})\n"
        preview += f"Columns per ring: {list(cols_per_ring.values())}\n\n"
        
        # Show first few rows as sample
        preview += "Sample data (first 3 rows):\n"
        for i, row in enumerate(data[:3]):
            preview += f"Ring {row['ring']}, Col {row['col']}: "
            preview += f"height={row['wave_height_mm']:.3f}mm, "
            preview += f"width={row['wave_width_mm']:.3f}mm\n"
            
        if len(data) > 3:
            preview += f"... and {len(data) - 3} more rows"
            
        if preview_text:
            preview_text.text = preview
        if status_text:
            status_text.text = f'Ready to process {total_rows} data points from {len(rings)} rings.'
        
        # Update calculated diameter
        diameter_input = adsk.core.ValueCommandInput.cast(inputs.itemById('diameter'))
        if diameter_input and data:
            # Calculate diameter from wave width
            avg_wave_width = sum(row['wave_width_mm'] for row in data) / len(data)
            cols_in_first_ring = len([row for row in data if row['ring'] == rings[0]])
            calculated_diameter = (avg_wave_width * cols_in_first_ring) / math.pi
            diameter_input.value = calculated_diameter / 10  # Convert mm to cm for Fusion
        
    except Exception as e:
        if preview_text:
            preview_text.text = f'Error reading file: {str(e)}'
        if status_text:
            status_text.text = f'Error: {str(e)}'
        futil.log(f'Error in update_data_preview: {traceback.format_exc()}')

def read_excel_data(file_path):
    """Read Excel data from WaveInputsByColumn sheet"""
    try:
        # For now, provide a simple CSV alternative or require manual data entry
        # This is a placeholder that would work with CSV data
        
        # Check if it's a CSV file instead
        if file_path.lower().endswith('.csv'):
            return read_csv_data(file_path)
        
        # For Excel files, we'll need openpyxl or pandas
        # Try to import openpyxl
        try:
            import openpyxl  # type: ignore
            
            wb = openpyxl.load_workbook(file_path, data_only=True)
            
            if 'WaveInputsByColumn' not in wb.sheetnames:
                raise Exception("Sheet 'WaveInputsByColumn' not found in Excel file.")
                
            ws = wb['WaveInputsByColumn']
            
            # Get header row
            headers = []
            for cell in ws[1]:
                if cell.value:
                    headers.append(str(cell.value).strip())
                    
            # Read data rows
            data = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] is not None:  # Skip empty rows
                    row_data = {}
                    for i, header in enumerate(headers):
                        if i < len(row) and row[i] is not None:
                            value = row[i]
                            try:
                                # Convert to appropriate type
                                if header in ['ring', 'col']:
                                    row_data[header] = int(float(str(value)))
                                else:
                                    row_data[header] = float(str(value))
                            except (ValueError, TypeError):
                                row_data[header] = str(value)
                        else:
                            row_data[header] = 0.0 if header not in ['ring', 'col'] else 0
                    data.append(row_data)
            
            wb.close()
            return data
            
        except ImportError:
            # If openpyxl is not available, suggest alternative
            raise Exception("openpyxl library not found. Please install openpyxl or save your Excel file as CSV format.")
            
    except Exception as e:
        futil.log(f'Error reading Excel file: {traceback.format_exc()}')
        raise

def read_csv_data(file_path):
    """Read CSV data as alternative to Excel"""
    try:
        import csv
        data = []
        
        with open(file_path, 'r', newline='', encoding='utf-8') as csvfile:
            reader = csv.DictReader(csvfile)
            
            for row in reader:
                row_data = {}
                for key, value in row.items():
                    try:
                        if key in ['ring', 'col']:
                            row_data[key] = int(float(value))
                        else:
                            row_data[key] = float(value)
                    except (ValueError, TypeError):
                        row_data[key] = 0.0 if key not in ['ring', 'col'] else 0
                data.append(row_data)
                
        return data
        
    except Exception as e:
        futil.log(f'Error reading CSV file: {traceback.format_exc()}')
        raise

def process_excel_file(file_path, diameter_mm, draw_construction=True, draw_chords=True, create_points=False):
    """Process the Excel file and create the stent frame sketch"""
    try:
        # Read Excel data
        data = read_excel_data(file_path)
        
        if not data:
            raise Exception("No data found in Excel file")
            
        # Get Fusion objects
        app = adsk.core.Application.get()
        ui = app.userInterface
        design = adsk.fusion.Design.cast(app.activeProduct)
        
        if not design:
            raise Exception("Please create a new design or open an existing one before running this command.")
            
        root = design.rootComponent
        
        # Create new sketch
        sketch = root.sketches.add(root.xYConstructionPlane)
        
        # Analyze data structure
        rings = sorted(set(row['ring'] for row in data))
        
        # Calculate stent dimensions
        first_ring_data = [row for row in data if row['ring'] == rings[0]]
        cols_per_ring = len(first_ring_data)
        
        # Calculate total length from ring heights and gaps
        total_length_mm = 0
        ring_positions = {}
        current_y = 0
        
        for ring_num in rings:
            ring_data = [row for row in data if row['ring'] == ring_num]
            ring_height = max(row['wave_height_mm'] for row in ring_data)
            
            ring_positions[ring_num] = {
                'center_y': current_y + ring_height / 2,
                'start_y': current_y,
                'end_y': current_y + ring_height,
                'height': ring_height
            }
            
            current_y += ring_height
            
            # Add gap after ring (except for last ring)
            if ring_num < max(rings):
                # Use gap_below from current ring data
                gap_below = max(row.get('gap_below_mm', 0) for row in ring_data)
                current_y += gap_below
                
        total_length_mm = current_y
        
        # Calculate width from diameter
        width_mm = diameter_mm * math.pi
        
        # Set sketch name
        sketch.name = f'Stent Frame from Excel - {len(rings)} rings, {cols_per_ring} cols'
        
        lines = sketch.sketchCurves.sketchLines
        
        # Convert mm to cm for Fusion API
        def mm_to_cm(x):
            return x * 0.1
            
        # Draw border
        if draw_construction:
            # Left border
            lines.addByTwoPoints(
                adsk.core.Point3D.create(0, 0, 0),
                adsk.core.Point3D.create(0, mm_to_cm(total_length_mm), 0)
            ).isConstruction = True
            
            # Right border
            lines.addByTwoPoints(
                adsk.core.Point3D.create(mm_to_cm(width_mm), 0, 0),
                adsk.core.Point3D.create(mm_to_cm(width_mm), mm_to_cm(total_length_mm), 0)
            ).isConstruction = True
            
            # Top border
            lines.addByTwoPoints(
                adsk.core.Point3D.create(0, mm_to_cm(total_length_mm), 0),
                adsk.core.Point3D.create(mm_to_cm(width_mm), mm_to_cm(total_length_mm), 0)
            ).isConstruction = True
            
            # Bottom border
            lines.addByTwoPoints(
                adsk.core.Point3D.create(0, 0, 0),
                adsk.core.Point3D.create(mm_to_cm(width_mm), 0, 0)
            ).isConstruction = True
            
        # Draw ring boundaries
        if draw_construction:
            for ring_num, ring_info in ring_positions.items():
                # Ring start line
                lines.addByTwoPoints(
                    adsk.core.Point3D.create(0, mm_to_cm(ring_info['start_y']), 0),
                    adsk.core.Point3D.create(mm_to_cm(width_mm), mm_to_cm(ring_info['start_y']), 0)
                ).isConstruction = True
                
                # Ring end line
                lines.addByTwoPoints(
                    adsk.core.Point3D.create(0, mm_to_cm(ring_info['end_y']), 0),
                    adsk.core.Point3D.create(mm_to_cm(width_mm), mm_to_cm(ring_info['end_y']), 0)
                ).isConstruction = True
                
        # Draw column boundaries
        if draw_construction:
            col_spacing = width_mm / cols_per_ring
            for col in range(1, cols_per_ring):
                x_pos = col * col_spacing
                lines.addByTwoPoints(
                    adsk.core.Point3D.create(mm_to_cm(x_pos), 0, 0),
                    adsk.core.Point3D.create(mm_to_cm(x_pos), mm_to_cm(total_length_mm), 0)
                ).isConstruction = True
                
        # Draw chord lines based on Excel data
        if draw_chords:
            col_spacing = width_mm / cols_per_ring
            
            for row in data:
                ring_num = row['ring']
                col_num = row['col']
                
                if ring_num in ring_positions:
                    ring_info = ring_positions[ring_num]
                    
                    # Calculate column center position
                    col_center_x = (col_num + 0.5) * col_spacing
                    
                    # Get chord and sagitta data
                    upper_chord = row.get('upper_chord_center_mm', 0)
                    upper_sagitta = row.get('upper_sagitta_center_mm', 0)
                    lower_chord = row.get('lower_chord_center_mm', 0)
                    lower_sagitta = row.get('lower_sagitta_center_mm', 0)
                    
                    # Draw upper chord line
                    if upper_chord > 0:
                        chord_half_length = upper_chord / 2
                        chord_y = ring_info['center_y'] + ring_info['height']/2 - upper_sagitta
                        
                        lines.addByTwoPoints(
                            adsk.core.Point3D.create(mm_to_cm(col_center_x - chord_half_length), mm_to_cm(chord_y), 0),
                            adsk.core.Point3D.create(mm_to_cm(col_center_x + chord_half_length), mm_to_cm(chord_y), 0)
                        ).isConstruction = True
                        
                    # Draw lower chord line
                    if lower_chord > 0:
                        chord_half_length = lower_chord / 2
                        chord_y = ring_info['center_y'] - ring_info['height']/2 + lower_sagitta
                        
                        lines.addByTwoPoints(
                            adsk.core.Point3D.create(mm_to_cm(col_center_x - chord_half_length), mm_to_cm(chord_y), 0),
                            adsk.core.Point3D.create(mm_to_cm(col_center_x + chord_half_length), mm_to_cm(chord_y), 0)
                        ).isConstruction = True
                        
        # Create points at intersections if requested
        if create_points:
            col_spacing = width_mm / cols_per_ring
            
            for ring_num, ring_info in ring_positions.items():
                for col in range(cols_per_ring + 1):
                    x_pos = col * col_spacing
                    
                    # Points at ring boundaries
                    sketch.sketchPoints.add(adsk.core.Point3D.create(mm_to_cm(x_pos), mm_to_cm(ring_info['start_y']), 0))
                    sketch.sketchPoints.add(adsk.core.Point3D.create(mm_to_cm(x_pos), mm_to_cm(ring_info['end_y']), 0))
                    
        # Show summary
        rings_count = len(rings)
        data_points = len(data)
        
        ui.messageBox(
            f'Stent frame created from Excel data!\n\n'
            f'• File: {os.path.basename(file_path)}\n'
            f'• Data points processed: {data_points}\n'
            f'• Rings: {rings_count} (rings {min(rings)} to {max(rings)})\n'
            f'• Columns per ring: {cols_per_ring}\n'
            f'• Calculated diameter: {diameter_mm:.2f} mm\n'
            f'• Calculated length: {total_length_mm:.2f} mm\n'
            f'• Width (circumference): {width_mm:.2f} mm\n'
            f'• Construction lines: {"Yes" if draw_construction else "No"}\n'
            f'• Chord lines: {"Yes" if draw_chords else "No"}\n'
            f'• Sketch points: {"Yes" if create_points else "No"}'
        )
        
    except Exception as e:
        futil.log(f'Error in process_excel_file: {traceback.format_exc()}')
        raise

# This event handler is called when the user interacts with any of the inputs in the dialog
# which allows you to verify that all of the inputs are valid and enables the OK button.
def command_validate_input(args: adsk.core.ValidateInputsEventArgs):
    # General logging for debug.
    futil.log(f'{CMD_NAME} Validate Input Event')

    inputs = args.inputs
    
    # Check if file path is provided and exists
    file_path_input = adsk.core.StringValueCommandInput.cast(inputs.itemById('file_path'))
    
    if file_path_input and file_path_input.value and os.path.exists(file_path_input.value):
        args.areInputsValid = True
    else:
        args.areInputsValid = False

# This event handler is called when the command terminates.
def command_destroy(args: adsk.core.CommandEventArgs):
    # General logging for debug.
    futil.log(f'{CMD_NAME} Command Destroy Event')

    global local_handlers
    local_handlers = []
